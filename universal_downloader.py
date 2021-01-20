"""Download documents and store properties."""

import os
import csv
import logging

from requests.sessions import Session
from PyPDF2 import PdfFileReader
from openpyxl import load_workbook

INPUT_CSV_FILENAME = 'test_doc_urls.csv'

OUTPUT_CSV_FILENAME = 'doc_metadata.csv'


def print_data(data: str, filename: str) -> True:
    """
    Print utilities

    :param data: dictionary with file properties
    :param filename: name of the document

    :returns: True

    """
    txt = f"""
    Information about {filename}: 

    Author: {data["Author"]}
    Title: {data["Title"]}
    Subject: {data["Subject"]}
    Keywords: {data["Keywords"]}

    """
    print(txt)
    return True


def empty_dict(doc_id: str, web_address: str) -> dict:
    """
    Create an empty dictionary

    :param doc_id: ID of file
    :param web_address: address of file

    :returns: empty dictionary

    """
    return {
        'DOC ID': doc_id,
        'Website URL': web_address,
        'Author': '',
        'Title': '',
        'Subject': '',
        'Keywords': '',
        'Number of pages': '',
        'Error': 'None',
        'Row': 'None',
    }


def write_temp_file(filename: str, file_content: str) -> True:
    """
    Write downloaded file into a temporary file

    :param filename: name of temporary file
    :param file_content: file content in binary format

    :returns: True

    """
    f = open(filename, "wb")
    f.write(file_content)
    f.close()
    return True


def extract_pdf_info(temp_pdf: str, doc_id: str, 
                      web_address: str, file_content: str) -> dict:
    """
    Extract and return information from PDF file

    :param temp_pdf: downloaded file
    :param doc_id: ID of file
    :param web_address: address of file
    :param file_content: encoded file content

    :returns: dictionary with file properties

    """
    write_temp_file(temp_pdf, file_content)
    data = empty_dict(doc_id, web_address)
    with open(temp_pdf, 'rb') as f:
        try:
            pdf = PdfFileReader(f)  
            if pdf.isEncrypted:
                pdf.decrypt("")
        except Exception as err:
            print(str(err))
            data['Error'] = f"From PdfFileReader {str(err)}"
            return data
        information = pdf.getDocumentInfo()
        try:
            keywords = pdf.getXmpMetadata()
        except Exception as err:
            print(str(err))
            data['Error'] = f"From getXmpMetadata {str(err)}"
            return data
        number_of_pages = pdf.getNumPages()
    try:
        author = information.author
    except AttributeError:
        author = "None"
    try:
        title = information.title
    except AttributeError:
        title = "None"
    try:
        subject = information.subject
    except AttributeError:
        subject = "None"
    try:
        pdf_keywords = keywords.pdf_keywords
    except AttributeError:
        pdf_keywords = "None"

    data["Author"] = author
    data["Title"] = title
    data["Subject"] = subject
    data["Keywords"] = pdf_keywords
    data["Number of pages"] = number_of_pages
    return data 


def extract_xlsx_info(temp_xlsx: str, doc_id: str, 
                      web_address: str, file_content: str):
    """
    Extract and return information from Excel file
    
    :param temp_xlsx: downloaded file
    :param doc_id: ID of file
    :param web_address: address of file
    :param file_content: encoded file content

    :returns: dictionary with file properties

    """
    write_temp_file(temp_xlsx, file_content)
    data = empty_dict(doc_id, web_address)
    with open(temp_xlsx, 'rb') as f:
        try:
            wb = load_workbook(f)  
        except Exception as err:
            print(str(err))
            data['Error'] = f"From load_workbook {str(err)}"
            return data
        prop = wb.properties

    try:
        creator = prop.creator
        if creator == "openpyxl":
            creator = "Empty"
    except AttributeError:
        creator = "None"
    try:
        title = prop.title
        if not title:
            title = "Empty"
    except AttributeError:
        title = "None"
    try:
        subject = prop.subject
    except AttributeError:
        subject = "None"
    try:
        keywords = prop.keywords
    except AttributeError:
        keywords = "None"

    data["Author"] = creator
    data["Title"] = title
    data["Subject"] = subject
    data["Keywords"] = keywords
    return data


def handle_masterfile(writer: object) -> True:
    """
    Iterate over every row in csv file and call appropriate handling function

    :param writer: function used to write properties into output csv file 

    :returns: True

    """
    # For each URL:
    print('open master csv file')
    session = Session()
    with open(INPUT_CSV_FILENAME, encoding="utf-8") as doc_list:
        csv_reader = csv.reader(doc_list)
        next(csv_reader)  # Skip first row of csv file
        for counter, row in enumerate(csv_reader):
            web_address = row[8]
            file_type = row[2]
            doc_id = row[0]
            expected_author = row[5]             
            print(f'Working on: {web_address}')
            s = session.get(web_address)
            if file_type.upper() == 'EXCEL':
                logging.info(web_address)
                data = extract_xlsx_info("temp.xlsx", doc_id, web_address, s.content)   
            elif file_type.upper() == 'PDF':
                logging.info(web_address)
                data = extract_pdf_info("temp.pdf", doc_id, web_address, s.content)   
            else:
                data = empty_dict(doc_id, web_address)
                data['Error']: f"NOT PDF OR EXCEL [{file_type}]"
                    
            data['Author'] = f"{expected_author} - ({data['Author']})"        
            data["Row"] = counter
            print_data(data, os.path.basename(web_address))
            writer.writerow(data)  # write data into csv file
            print("...done")
    return True


def handle_outputfile():
    """Prepare output file for writing."""   
    logging.info('Open output file')
    print('open output file')
    with open(OUTPUT_CSV_FILENAME, 'w', newline='', encoding="utf-8") as csvfile:
        data_dummy = empty_dict("None", "None")
        fieldnames = list(data_dummy.keys())
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        handle_masterfile(writer)
    return True

 
def main():
    """Call main function."""
    logging.basicConfig(filename = 'universal_downloader.log', level = logging.INFO)
    logging.info('Start universal downloader')
    handle_outputfile()
    return True
                                
                
if __name__ == "__main__":
    main()
    print("The End")

